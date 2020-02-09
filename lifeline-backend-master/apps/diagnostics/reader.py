from openpyxl import load_workbook
from django.utils.translation import ugettext_lazy as _

from apps.diagnostics.models import Diagnose, HealthScreening


class XslxReader:
    content_error = _('Wrong format')
    no_data_error = _('No data was parsed from file')

    def __init__(self, diagnostic):
        self._errors = []
        if diagnostic.file:
            self.diagnostic = diagnostic
            self.diagnostic.save(is_uploaded=False)
            self._wb = load_workbook(diagnostic.file)
        else:
            self._errors.append('File was not uploaded to server')

    def proceed_diagnoses(self):
        results = []
        if not self._errors:
            try:
                table = self._wb[self._wb.sheetnames[2]].rows
                iterator = next(table)
                french_map = self._get_table_map(0)
                dutch_map = self._get_table_map(1)
                if iterator[0].value != 'conceptId' or iterator[1].value != 'term':
                    self._errors.append(_('File has wrong content'))
                    return
                for row in table:
                    results.append(Diagnose(
                        concept_id=row[0].value,
                        term_en=row[1].value,
                        term_fr=french_map.get(row[0].value),
                        term_nl=dutch_map.get(row[0].value)
                    ))
                if not results:
                    self._errors.append(self.no_data_error)
                if not self._errors:
                    Diagnose.objects.all().delete()
                    Diagnose.objects.bulk_create(results)
                    self.diagnostic.errors = None
                    self.diagnostic.save(is_uploaded=True)
            except IndexError:
                self._errors.append(self.content_error)
        self._proceed_errors()

    def _get_table_map(self, index):
        rows = self._wb[self._wb.sheetnames[index]].rows
        iterator = next(rows)
        if iterator[0].value != 'conceptId' or iterator[1].value != 'term' or len(iterator) != 2:
            self._errors.append(_('File has wrong content'))
            return {}
        return {str(row[0].value): row[1].value for row in rows}

    def proceed_health_screening(self):
        results = []
        if not self._errors:
            try:
                data_fr = self._wb[self._wb.sheetnames[0]].rows
                headers_fr = next(data_fr)
                data = self._wb[self._wb.sheetnames[1]].rows
                headers = next(data)
                data_nl = self._wb[self._wb.sheetnames[2]].rows
                headers_nl = next(data_nl)

                if headers[0].value != 'Particular monitoring type' \
                        or headers_nl[0].value != 'Bijzonder toezicht type'\
                        or headers_fr[0].value != 'Type de surveillance particulière':
                    self._errors.append(self.content_error)
                    return
                for index, row in enumerate(data):
                    results.append(HealthScreening(
                        term_en=row[0].value,
                        term_fr=next(data_fr)[0].value,
                        term_nl=next(data_nl)[0].value
                    ))
                if not results:
                    self._errors.append(self.no_data_error)
                if not self._errors:
                    HealthScreening.objects.all().delete()
                    HealthScreening.objects.bulk_create(results)
                    self.diagnostic.save(is_uploaded=True)
            except IndexError:
                self._errors.append(self.content_error)

    def get_error(self):
        return self._errors[0] if self._errors else None

    def _proceed_errors(self):
        error = self.get_error()
        if error:
            self.diagnostic.errors = error
            self.diagnostic.save(is_uploaded=False)
